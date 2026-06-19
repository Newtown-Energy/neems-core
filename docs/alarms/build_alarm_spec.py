#!/usr/bin/env python3
"""Generate a structured JSON spec from the client alarm/SLD spreadsheet.

Reads the client alarm/SLD source spreadsheet (sheets: Digitals, Analogs) and
emits ``newtown-alarms.json`` — a faithful, normalized capture of every row plus
deterministically *derived* fields (snake_case names, Modbus register/bit
positions) that downstream implementation can rely on.

The source spreadsheet is kept OUTSIDE the repository (it carries vendor/site
names that must not be committed). Point this script at it with the ALARM_XLSX
environment variable, or place a single ``.xlsx`` in the project root (the
directory that contains neems-core).

Run:

    ALARM_XLSX=/path/to/source.xlsx python3 neems-core/docs/alarms/build_alarm_spec.py

Source-of-truth notes:
  * Every value taken verbatim from the spreadsheet lives under "source"-style
    keys (name, zone, sld.*_raw, threshold_raw, mouseover, ...).
  * Every value we computed is grouped/labeled and marked with "_derived": true
    (suggested_code_name, modbus, severity_signals, threshold parse).
"""

import json
import os
import re
import sys
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
# Project root = the directory containing neems-core (.../<root>/neems-core/docs/alarms).
PROJECT_ROOT = SCRIPT_DIR.parents[2]
OUT = SCRIPT_DIR / "newtown-alarms.json"


def _find_xlsx():
    """Locate the source spreadsheet, which lives outside the repo. Prefer the
    ALARM_XLSX env var; otherwise use the sole .xlsx in the project root."""
    env = os.environ.get("ALARM_XLSX")
    if env:
        return Path(env)
    candidates = sorted(PROJECT_ROOT.glob("*.xlsx"))
    if len(candidates) == 1:
        return candidates[0]
    return None


XLSX = _find_xlsx()

# Neutral provenance label recorded in the output. We deliberately do NOT embed
# the real filename, which carries vendor/site names that must not be committed.
SOURCE_FILE = "client alarm & SLD source spreadsheet (kept outside the repo)"


# --------------------------------------------------------------------------- #
# Reference data
# --------------------------------------------------------------------------- #

# Spreadsheet zone label -> canonical code zone + React SLD component id.
# Digitals use labels like "MP1A_digital"; Analogs use "MP-1A". Both map to the
# same canonical zone.
ZONE_MAP = {
    "Newtown": ("Site", "site"),
    "Breaker Relay SEL-451": ("BreakerRelay", "breaker-main"),
    "Meter 1 SEL735": ("Meter", "meter-main"),
    "Transformer 1": ("Transformer1", "transformer-1"),
    "Transformer 2": ("Transformer2", "transformer-2"),
    "RTAC": ("Rtac", "rtac"),
    "FACP": ("Facp", "fire-alarm-panel"),
    "TeslaSiteController": ("TeslaSiteController", "tesla-site-controller"),
    "MP1A_digital": ("Mp1a", "megapack-1a"),
    "MP1B_digital": ("Mp1b", "megapack-1b"),
    "MP1C_digital": ("Mp1c", "megapack-1c"),
    "MP2A_digital": ("Mp2a", "megapack-2a"),
    "MP2B_digital": ("Mp2b", "megapack-2b"),
    "MP2C_digital": ("Mp2c", "megapack-2c"),
    "MP-1A": ("Mp1a", "megapack-1a"),
    "MP-1B": ("Mp1b", "megapack-1b"),
    "MP-1C": ("Mp1c", "megapack-1c"),
    "MP-2A": ("Mp2a", "megapack-2a"),
    "MP-2B": ("Mp2b", "megapack-2b"),
    "MP-2C": ("Mp2c", "megapack-2c"),
}

# Digital alarm-register layout. This is a MANUAL MIRROR of the layout the
# backend already encodes in neems-data/src/rtac/alarm_definitions.rs (and
# protocol.rs: ALARM_REGISTER_COUNT = 22, alarm block starts at address 8). It
# is duplicated here only so this one-shot generator can run without a Rust
# toolchain — it is NOT a separate source of truth. If the Rust layout changes,
# update this table to match (or the derived `modbus` fields will silently drift).
#
# Each zone block is allocated a base register and a number of contiguous 16-bit
# registers; bit = (alarm_num - base_num) % 16, register_index = base_reg +
# (alarm_num - base_num) // 16. Entries whose offset exceeds the allocation (only
# stray trailing "reserved" numbers do) get a null Modbus mapping and a
# data-quality note.
#
# The first alarm register sits at Modbus holding-register address 8 (after the
# 8 status registers), so register_address = 8 + register_index.
ALARM_REGISTER_BASE_ADDRESS = 8
DIGITAL_BLOCKS = [
    # (canonical_zone, base_alarm_num, base_register_index, registers_allocated)
    ("Site", 1, 0, 1),
    ("BreakerRelay", 101, 1, 2),
    ("Meter", 201, 3, 1),
    ("Transformer1", 301, 4, 1),
    ("Transformer2", 311, 5, 1),
    ("Rtac", 321, 6, 1),
    ("Facp", 401, 7, 2),
    ("TeslaSiteController", 501, 9, 1),
    ("Mp1a", 601, 10, 2),
    ("Mp1b", 631, 12, 2),
    ("Mp1c", 661, 14, 2),
    ("Mp2a", 691, 16, 2),
    ("Mp2b", 721, 18, 2),
    ("Mp2c", 751, 20, 2),
]


def infer_digital_zone(alarm_num):
    """Infer the canonical zone for a digital alarm number from its numbering
    block. Used to backfill the zone for blank-zone "reserved" rows whose
    spreadsheet zone cell was left empty."""
    bases = sorted(((base, zname) for zname, base, _, _ in DIGITAL_BLOCKS))
    inferred = None
    for base, zname in bases:
        if alarm_num >= base:
            inferred = zname
        else:
            break
    return inferred


def digital_modbus(zone_code, alarm_num):
    """Return (modbus_dict or None, note or None) for a digital alarm bit."""
    for zname, base_num, base_reg, regs in DIGITAL_BLOCKS:
        if zname != zone_code:
            continue
        offset = alarm_num - base_num
        if offset < 0:
            continue
        reg = base_reg + offset // 16
        bit = offset % 16
        if offset // 16 >= regs:
            return None, (
                f"alarm {alarm_num} falls outside the {regs}-register allocation "
                f"for zone {zone_code}; no Modbus bit assigned"
            )
        return {
            "register_index": reg,
            "register_address": ALARM_REGISTER_BASE_ADDRESS + reg,
            "bit": bit,
            "_derived": True,
        }, None
    return None, None


# --------------------------------------------------------------------------- #
# Parsing helpers
# --------------------------------------------------------------------------- #

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def snake(name):
    """Mechanical snake_case suggestion from a human alarm name."""
    if not name:
        return None
    s = name.strip().lower()
    s = s.replace("/", " ")
    s = re.sub(r"[()\[\]\"']", " ", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or None


def strip_quotes(v):
    if v is None:
        return None
    s = str(v).strip()
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1]
    return s.strip() or None


def split_sld_objects(v):
    if v is None:
        return []
    return [tok.strip() for tok in str(v).split(",") if tok.strip()]


def parse_sld_changes(raw):
    """Best-effort structured split of an "SLD Change" cell.

    e.g. "Main obj: Red, flashing; Border obj: blue, flash"
      -> [{target:"Main obj", instruction:"Red, flashing"},
          {target:"Border obj", instruction:"blue, flash"}]
    Segments without a "target:" prefix are kept as {target:null, instruction:...}.
    """
    if raw is None:
        return []
    out = []
    for seg in str(raw).split(";"):
        seg = seg.strip()
        if not seg:
            continue
        if ":" in seg:
            target, instr = seg.split(":", 1)
            out.append({"target": target.strip(), "instruction": instr.strip()})
        else:
            out.append({"target": None, "instruction": seg})
    return out


COLOR_RE = re.compile(r"\b(red|yellow|green|blue|orange|amber)\b", re.I)


def severity_signals(sld_change_raw, is_fire, availability_impact):
    """Capture the raw severity signals the spreadsheet encodes (color of the
    primary SLD object, whether it flashes, fire flag, availability impact).
    Final numeric levels are intentionally left to the implementation step."""
    primary_color = None
    flashing = None
    if sld_change_raw:
        # Color of the "Main obj" if present, else first color mentioned.
        main_seg = None
        for seg in str(sld_change_raw).split(";"):
            if "main" in seg.lower():
                main_seg = seg
                break
        scan = main_seg if main_seg else str(sld_change_raw)
        m = COLOR_RE.search(scan)
        if m:
            primary_color = m.group(1).lower()
        flashing = bool(re.search(r"\bflash", str(sld_change_raw), re.I))
    return {
        "primary_color": primary_color,
        "flashing": flashing,
        "is_fire": bool(is_fire),
        "availability_impact": availability_impact,
        "_derived": True,
    }


def parse_threshold(raw):
    if raw is None:
        return None
    m = re.match(r"^\s*(-?\d+(?:\.\d+)?)\s*([A-Za-z%]+)?\s*$", str(raw))
    if not m:
        return None
    val = float(m.group(1))
    if val.is_integer():
        val = int(val)
    return {"value": val, "unit": m.group(2) or None, "_derived": True}


def norm_avail(v):
    """Normalize the availability-impact column to a small enum-ish string."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s == "site offline":
        return "site_offline"
    if s == "mp offline":
        return "mp_offline"
    return s


# --------------------------------------------------------------------------- #
# Row -> entry
# --------------------------------------------------------------------------- #

def rows(ws):
    for r in range(1, ws.max_row + 1):
        yield [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]


def build_digital(row):
    num = row[0]
    if not isinstance(num, int):
        return None
    zone_raw = clean(row[1])
    name = clean(row[2])
    pt = row[3] if isinstance(row[3], int) else None
    sld_obj_raw = clean(row[4])
    sld_change_raw = clean(row[5])
    avail = norm_avail(clean(row[6]))
    mouseover = strip_quotes(row[7])
    is_fire = clean(row[8]) is not None and str(row[8]).strip().upper().startswith("Y")

    reserved = name is None or "reserved" in name.lower() or name == "[RESERVED]"
    zone_code, sld_component = ZONE_MAP.get(zone_raw, (None, None))

    # Backfill zone for blank-zone reserved rows from the numbering block.
    zone_inferred = False
    if zone_code is None:
        zone_code = infer_digital_zone(num)
        if zone_code is not None:
            zone_inferred = True
            _, sld_component = next(
                (c, comp) for c, comp, _ in _zone_reference() if c == zone_code
            )

    modbus, note = (None, None)
    if zone_code:
        modbus, note = digital_modbus(zone_code, num)

    entry = {
        "alarm_num": num,
        "category": "digital",
        "zone_raw": zone_raw,
        "zone": zone_code,
        "zone_inferred": zone_inferred,
        "sld_component_id": sld_component,
        "name": name,
        "suggested_code_name": snake(name),
        "reserved": reserved,
        "pt_number": pt,
        "sld": {
            "related_objects_raw": sld_obj_raw,
            "related_objects": split_sld_objects(sld_obj_raw),
            "change_raw": sld_change_raw,
            "changes": parse_sld_changes(sld_change_raw),
        },
        "availability_impact": avail,
        "mouseover": mouseover,
        "is_fire": is_fire,
        "severity_signals": severity_signals(sld_change_raw, is_fire, avail),
        "modbus": modbus,
    }
    return entry, note


def build_analog(row):
    num = row[0]
    if not isinstance(num, int):
        return None
    zone_raw = clean(row[1])
    name = clean(row[2])
    pt = row[3] if isinstance(row[3], int) else None
    sld_obj_raw = clean(row[4])
    sld_change_raw = clean(row[5])
    avail = norm_avail(clean(row[6]))
    threshold_raw = clean(row[7])
    # row[8] is a duplicate "MP or Site Availability" header column (always empty)
    mouseover = strip_quotes(row[9])
    is_fire = clean(row[10]) is not None and str(row[10]).strip().upper().startswith("Y")
    alarm_levels = clean(row[11])

    reserved = name is not None and (
        name.lower().startswith("ai_spare") or name.lower().startswith("di_spare")
    )
    zone_code, sld_component = ZONE_MAP.get(zone_raw, (None, None))

    entry = {
        "alarm_num": num,
        "category": "analog",
        "zone_raw": zone_raw,
        "zone": zone_code,
        "sld_component_id": sld_component,
        "name": name,
        "suggested_code_name": snake(name),
        "spare": reserved,
        "pt_number": pt,
        "sld": {
            "related_objects_raw": sld_obj_raw,
            "related_objects": split_sld_objects(sld_obj_raw),
            "change_raw": sld_change_raw,
            "changes": parse_sld_changes(sld_change_raw),
        },
        "availability_impact": avail,
        "threshold_raw": threshold_raw,
        "threshold": parse_threshold(threshold_raw),
        "mouseover": mouseover,
        "is_fire": is_fire,
        "alarm_levels": alarm_levels,
    }
    return entry


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main():
    if XLSX is None or not XLSX.exists():
        sys.exit(
            "source spreadsheet not found. Set ALARM_XLSX=/path/to/source.xlsx "
            f"or place exactly one .xlsx in {PROJECT_ROOT}."
        )
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    digital_alarms = []
    dq_issues = []
    for row in rows(wb["Digitals"]):
        built = build_digital(row)
        if not built:
            continue
        entry, note = built
        digital_alarms.append(entry)
        if note:
            dq_issues.append({"alarm_num": entry["alarm_num"], "category": "digital", "issue": note})
        if entry["zone"] is None and entry["zone_raw"] is not None:
            dq_issues.append({
                "alarm_num": entry["alarm_num"], "category": "digital",
                "issue": f"unmapped zone label {entry['zone_raw']!r}",
            })

    analog_points = []
    for row in rows(wb["Analogs"]):
        entry = build_analog(row)
        if not entry:
            continue
        analog_points.append(entry)
        if entry["zone"] is None:
            dq_issues.append({
                "alarm_num": entry["alarm_num"], "category": "analog",
                "issue": f"unmapped zone label {entry['zone_raw']!r}",
            })

    # Detect spreadsheet data-entry anomalies: any Megapack block whose point
    # names diverge from the MP-1A reference template (by offset for analogs, by
    # pt_number for digitals). This surfaces the known copy/paste mislabelings
    # (e.g. "ac_voltage_phaseA" where the template has "inverter_phaseA_current").
    MP_ZONES = ["Mp1b", "Mp1c", "Mp2a", "Mp2b", "Mp2c"]
    mp1a_analog_names = [e["name"] for e in analog_points if e["zone"] == "Mp1a"]
    for zcode in MP_ZONES:
        names = [e["name"] for e in analog_points if e["zone"] == zcode]
        nums = [e["alarm_num"] for e in analog_points if e["zone"] == zcode]
        # zip() would silently truncate to the shorter list, hiding points that
        # exist in one block but not the other; flag the length divergence too.
        if len(names) != len(mp1a_analog_names):
            dq_issues.append({
                "alarm_num": nums[0] if nums else None, "category": "analog",
                "issue": (
                    f"{zcode} has {len(names)} analog points but the MP-1A "
                    f"template has {len(mp1a_analog_names)}; only the overlapping "
                    f"prefix was checked for name mismatches"
                ),
            })
        for off, (ref, got) in enumerate(zip(mp1a_analog_names, names)):
            if ref != got:
                dq_issues.append({
                    "alarm_num": nums[off], "category": "analog",
                    "issue": (
                        f"MP analog point name mismatch at block offset {off}: "
                        f"{zcode} has {got!r} but MP-1A template has {ref!r} "
                        f"(likely a spreadsheet copy/paste error)"
                    ),
                })

    mp1a_digital = sorted(
        (e for e in digital_alarms if e["zone"] == "Mp1a"),
        key=lambda e: e["pt_number"] or 0,
    )
    # Key by pt_number, but skip unnumbered points: a None key would collide
    # across every pt_number-less row and produce meaningless comparisons.
    mp1a_digital_names = {
        e["pt_number"]: e["name"] for e in mp1a_digital if e["pt_number"] is not None
    }
    for zcode in MP_ZONES:
        for e in (a for a in digital_alarms if a["zone"] == zcode):
            if e["pt_number"] is None:
                continue
            ref = mp1a_digital_names.get(e["pt_number"])
            if ref is not None and ref != e["name"]:
                dq_issues.append({
                    "alarm_num": e["alarm_num"], "category": "digital",
                    "issue": (
                        f"MP digital point name mismatch at pt {e['pt_number']}: "
                        f"{zcode} has {e['name']!r} but MP-1A template has {ref!r}"
                    ),
                })

    # Megapack templates (canonical 30-point patterns; MP-1A is the clean copy).
    mp_digital_template = [
        {"offset": e["pt_number"] - 1 if e["pt_number"] else None,
         "pt_number": e["pt_number"], "name": e["name"],
         "suggested_code_name": e["suggested_code_name"]}
        for e in digital_alarms if e["zone"] == "Mp1a"
    ]
    mp_analog_template = [
        {"offset": i, "name": e["name"], "suggested_code_name": e["suggested_code_name"],
         "threshold": e["threshold"], "is_fire": e["is_fire"],
         "related_objects": e["sld"]["related_objects"]}
        for i, e in enumerate(e for e in analog_points if e["zone"] == "Mp1a")
    ]

    # Distinct SLD object tokens across both sheets.
    sld_tokens = {}
    for e in digital_alarms + analog_points:
        for tok in e["sld"]["related_objects"]:
            sld_tokens[tok] = sld_tokens.get(tok, 0) + 1

    spec = {
        "metadata": {
            "source_file": SOURCE_FILE,
            "source_sheets": ["Digitals", "Analogs"],
            "description": (
                "Structured capture of the client alarm + SLD register spreadsheet "
                "for the Newtown BESS. Intended as the single source of truth for "
                "replacing alarms and data points across the backend (neems-core), "
                "frontend (neems-react), and Modbus simulation (neems-rtac-sim)."
            ),
            "conventions": {
                "alarm_num_namespacing": (
                    "alarm_num is unique WITHIN a category but NOT across categories: "
                    "e.g. 601 is a digital MP-1A status bit AND an analog MP-1A "
                    "measurement. Always key by (category, alarm_num)."
                ),
                "derived_fields": (
                    "Any object containing \"_derived\": true was computed here, not "
                    "taken from the spreadsheet (suggested_code_name, modbus, "
                    "severity_signals, threshold). Raw spreadsheet values use *_raw keys."
                ),
                "modbus": (
                    "Digital alarms are packed into 22 contiguous 16-bit holding "
                    "registers starting at address 8. register_index is 0-based within "
                    "that block; register_address = 8 + register_index. Analog points "
                    "have no register assignment in the spreadsheet (TBD)."
                ),
                "severity": (
                    "The spreadsheet encodes severity indirectly via SLD color "
                    "(red = critical/trip, yellow = warning), the IsFire flag, and the "
                    "availability impact. These raw signals are captured under "
                    "severity_signals; mapping them to a 1-5 level is left to "
                    "implementation (see README)."
                ),
            },
            "counts": {
                "digital_alarms": len(digital_alarms),
                "digital_named": sum(1 for e in digital_alarms if not e["reserved"]),
                "digital_reserved": sum(1 for e in digital_alarms if e["reserved"]),
                "analog_points": len(analog_points),
                "analog_spare": sum(1 for e in analog_points if e["spare"]),
                "data_quality_issues": len(dq_issues),
            },
        },
        "reference": {
            "zones": [
                {"spreadsheet_labels": labels, "zone": code, "sld_component_id": comp}
                for code, comp, labels in _zone_reference()
            ],
            "severity_levels": {
                "1": "Emergency — active fire, call 911, emergency shutdown",
                "2": "High — activate COF, contact SMEs, immediate response",
                "3": "Medium — inform management, contact SMEs",
                "4": "Low — operator troubleshooting, business-hours escalation",
                "5": "Informational / unclassified",
            },
            "availability_impacts": ["site_offline", "mp_offline"],
            "sld_change_color_legend": {
                "red": "electrical/controls fault or fire (primary object red, often flashing)",
                "yellow": "warning / degraded but operable",
                "green": "normal / open switch indication",
                "blue": "border indication: site not ready to operate/close in (controls problem)",
                "border_red": "fire / life-safety emergency (main pane border)",
            },
            "sld_object_tokens": [
                {"token": tok, "occurrences": n}
                for tok, n in sorted(sld_tokens.items(), key=lambda kv: (-kv[1], kv[0]))
            ],
        },
        "megapack_digital_template": {
            "applies_to": ["Mp1a", "Mp1b", "Mp1c", "Mp2a", "Mp2b", "Mp2c"],
            "note": "30 status bits per Megapack, DNP3/Modbus points 1-30 (pt_number = bit offset + 1).",
            "points": mp_digital_template,
        },
        "megapack_analog_template": {
            "applies_to": ["Mp1a", "Mp1b", "Mp1c", "Mp2a", "Mp2b", "Mp2c"],
            "note": (
                "30 analog measurements per Megapack, captured in spreadsheet order "
                "(offset 0-29). MP-1A is the clean copy; see data_quality_issues for "
                "mislabelings in the other blocks."
            ),
            "points": mp_analog_template,
        },
        "digital_alarms": digital_alarms,
        "analog_points": analog_points,
        "data_quality_issues": dq_issues,
    }

    OUT.write_text(json.dumps(spec, indent=2, ensure_ascii=False) + "\n")
    print(f"wrote {OUT}")
    print(f"  digital_alarms: {len(digital_alarms)} "
          f"(named {spec['metadata']['counts']['digital_named']}, "
          f"reserved {spec['metadata']['counts']['digital_reserved']})")
    print(f"  analog_points:  {len(analog_points)} "
          f"(spare {spec['metadata']['counts']['analog_spare']})")
    print(f"  data_quality_issues: {len(dq_issues)}")
    print(f"  mp1a analog names ({len(mp1a_analog_names)}): {mp1a_analog_names}")


def _zone_reference():
    """Group ZONE_MAP back into (code, component, [labels])."""
    by_code = {}
    for label, (code, comp) in ZONE_MAP.items():
        by_code.setdefault(code, (comp, []))
        by_code[code][1].append(label)
    order = ["Site", "BreakerRelay", "Meter", "Transformer1", "Transformer2",
             "Rtac", "Facp", "TeslaSiteController",
             "Mp1a", "Mp1b", "Mp1c", "Mp2a", "Mp2b", "Mp2c"]
    return [(code, by_code[code][0], by_code[code][1]) for code in order]


if __name__ == "__main__":
    main()
